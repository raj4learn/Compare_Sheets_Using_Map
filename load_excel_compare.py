import sys
# import pandas as pd
import openpyxl as xl
from global_functions import validate_file, print_break, print_highlight, xl_head_row_with_config_comp, usage, conform_exit
import logger as log
import readConfigFile as rc
from openpyxl.styles.colors import RED
from openpyxl.styles import Font, Color, colors
from openpyxl.styles import Color, PatternFill, Font, Border
import re
from map_data_reading import read_mapper
import traceback as err

g_config_d = {}
g_ignore_list_d = {}

f_nm = ""


# log.set_module("ProcessBugList.py")
# log.set_log_level("debug")

#################################################################################
# Calling Xlsx Input File
#################################################################################
def read_xls(isCaseignore, p_file_nm, p_src_pk, p_dest_pk, p_xls_src_cols, p_xls_dest_cols, p_ignore_list, p_Src_Sh, p_Dest_Sh, p_Out_Sh = "CompOut"):
    m_nm = f_nm + ""
    log.debug(m_nm, "Start Reading Excel File", p_file_nm)
    l_error_flg = False
    l_tmp_str = ""

    try:
        print_break(80)
        l_src_col_nm_c_indx_d = {}
        l_dest_col_nm_c_indx_d = {}

        # Where the header Row in the Sheet
        try:
            l_xls_src_header_row_num = int(g_config_d['SRC_HEADER_ROW_NUM'])
            l_xls_dest_header_row_num = int(g_config_d['DEST_HEADER_ROW_NUM'])
            l_xls_map_header_row_num = int(g_config_d['MAPPING_HEADER_ROW_NUM'])
        except:
            l_tmp_str = "Default"
            l_xls_src_header_row_num = 1
            l_xls_dest_header_row_num = 1
            l_xls_map_header_row_num = 1

        print(f"Source Header Row [{l_xls_src_header_row_num}] : Destination Header Row [{l_xls_dest_header_row_num}]")

        wb_obj = xl.load_workbook(p_file_nm, data_only=True)

        try:
            l_xls_src_cols = list()  # list(str(g_config_d['COMP_SRC_COLS']).split(","))
            l_xls_dest_cols = list()  # list(str(g_config_d['COMP_DEST_COLS']).split(","))
            l_ignore_list = list()
            l_config_columns = 0

            g_src_pk_cell, g_dest_pk_cell, l_xls_src_cols, l_xls_dest_cols, g_ignore_list_d = p_src_pk, p_dest_pk, p_xls_src_cols, p_xls_dest_cols, p_ignore_list

            print(f"read_xls: Source Sheet  :{p_Src_Sh}")
            print(f"read_xls: Source Columns:{l_xls_src_cols}")
            print(f"read_xls: Source Key    :{g_src_pk_cell}")

            print(f"read_xls: Dest Sheet    :{p_Dest_Sh}")
            print(f"read_xls: Dest Columns  :{l_xls_dest_cols}")
            print(f"read_xls: Dest Key      :{g_dest_pk_cell}")

            print(f"read_xls: ignore List   :{g_ignore_list_d}")
            print(f"read_xls: ignore Case For All :{isCaseignore}")

            print(f"read_xls: Source column and Destination column are matching {l_config_columns}, Length of dest config variable {l_xls_dest_cols.__len__()}")
        except:
            print(f"Error when reading Mapping data, Read from {p_file_nm}")
            raise()

        '''
        try:
            if input_map_file != ':EXCEL:':
                l_xls_key_col_nm = str(g_config_d['XLSX_KEY_COL'])
                g_src_pk_cell = l_xls_key_col_nm.split(":")[0]
                g_dest_pk_cell = l_xls_key_col_nm.split(":")[1]

            log.info(m_nm,
                     "Key columns in Source and Destination File:[" + g_src_pk_cell + "]-[" + g_dest_pk_cell + "]")
        except:
            log.error(m_nm, "Unable to Get the Primary Key for the Source and Destination")
            raise
        '''
        try:

            '''
            if "SRC_EXCEL_SHEET_NAME" in g_config_d.keys() and "DEST_EXCEL_SHEET_NAME" in g_config_d.keys():
                sheet_obj_1 = wb_obj[g_config_d['SRC_EXCEL_SHEET_NAME']]
                sheet_obj_2 = wb_obj[g_config_d['DEST_EXCEL_SHEET_NAME']]
            else:
                if input_map_file == ':EXCEL:':
                    print(f"First three sheet names in excel file: {wb_obj.worksheets[0]}:{wb_obj.worksheets[1]}:{wb_obj.worksheets[2]}")
                    sheet_obj_1 = wb_obj.worksheets[1]  # wb_obj[xlsx_sheet_names[0]]
                    sheet_obj_2 = wb_obj.worksheets[2]  # wb_obj[xlsx_sheet_names[1]]
                else:
                    print(f"First two sheet names in excel file: {wb_obj.worksheets[0]}:{wb_obj.worksheets[1]}")
                    sheet_obj_1 = wb_obj.worksheets[0]  # wb_obj[xlsx_sheet_names[0]]
                    sheet_obj_2 = wb_obj.worksheets[1]  # wb_obj[xlsx_sheet_names[1]]
            '''
            g_sheets = { val : key for key, val in enumerate(list(wb_obj.sheetnames)) }
            print(f"Sheets: {g_sheets}")
            if type(p_Src_Sh) == int:
                sheet_obj_1 = wb_obj.worksheets[p_Src_Sh]
            else:
                sheet_obj_1 = wb_obj[p_Src_Sh]

            if type(p_Dest_Sh) == int:
                sheet_obj_2 = wb_obj.worksheets[p_Dest_Sh]
            else:
                sheet_obj_2 = wb_obj[p_Dest_Sh]

        except:
            print(f"Error: Missing first two sheets in the excel file {p_file_nm} [{p_Src_Sh}]:[{p_Dest_Sh}]")
            raise

        try:
            sheet_obj_op = wb_obj.create_sheet(p_Out_Sh, (g_sheets[p_Dest_Sh] + 1))
        except:
            print(f"Unable to Create a New Sheet in the Excel {p_file_nm}")
            raise

        print_break(80)
        ###### Start Reading Column Names and Getting the Index
        l_max_cols_1 = sheet_obj_1.max_column
        l_max_cols_2 = sheet_obj_2.max_column

        for x in range(1, l_max_cols_1 + 1):
            key = str(sheet_obj_1.cell(row=l_xls_src_header_row_num, column=x).value)
            nkey = key.strip()
            l_src_col_nm_c_indx_d[nkey] = x

        print(f"Source sheet available Columns: {l_src_col_nm_c_indx_d}")

        for x in range(1, l_max_cols_2 + 1):
            key = str(sheet_obj_2.cell(row=l_xls_dest_header_row_num, column=x).value)
            nkey = key.strip()
            l_dest_col_nm_c_indx_d[nkey] = x

        print(f"Destination sheet available Columns: {l_dest_col_nm_c_indx_d}")
        #####


        #####
        l_ret_val = xl_head_row_with_config_comp(l_src_col_nm_c_indx_d, l_xls_src_cols)
        if l_ret_val == False:
            print(f"Not All Map Source columns are available in the Source Excel Sheet {p_file_nm}")
            raise

        l_ret_val = xl_head_row_with_config_comp(l_dest_col_nm_c_indx_d, l_xls_dest_cols)
        if l_ret_val == False:
            print(f"Not All Map Destination columns are available in the Destination Excel Sheet {p_file_nm}")
            raise

        ##### End Reading Column Names and Getting the Index
        print_break(80)

        ##### Start Reading Primary Key Rows into New Sheet
        l_max_rows_1 = sheet_obj_1.max_row
        l_max_rows_2 = sheet_obj_2.max_row

        print(f"Rows Count In Source: [{l_max_rows_1}] - Rows Count In Destination [{l_max_rows_2}]")

        op_rows_idx = 1
        pk_value_idx_op_d = {}
        for x in range(1, l_max_rows_1 + 1):
            if x > l_xls_src_header_row_num:
                key = [sheet_obj_1.cell(row=x, column=l_src_col_nm_c_indx_d[cell_pk_nm]).value for cell_pk_nm in g_src_pk_cell]
                l_tmp_key = "#".join(map(str, key))
                cln_key = re.sub(r"\s", "", l_tmp_key)
                pk_value_idx_op_d[cln_key] = op_rows_idx

            op_rows_idx += 1

        for x in range(1, l_max_rows_2 + 1):
            if x > l_xls_dest_header_row_num:
                key = [sheet_obj_2.cell(row=x, column=l_dest_col_nm_c_indx_d[cell_pk_nm]).value for cell_pk_nm in g_dest_pk_cell]
                l_tmp_key = "#".join(map(str, key))
                cln_key = re.sub(r"\s", "", l_tmp_key)
                if cln_key not in pk_value_idx_op_d:
                    pk_value_idx_op_d[cln_key] = op_rows_idx
                    op_rows_idx += 1

        # print(pk_value_idx_op_d)
        op_wt_col_at = 4

        # Reading Row Names and Getting the Index
        my_red = xl.styles.colors.Color(rgb='00FF8080')
        my_green = xl.styles.colors.Color(rgb='00CCFFCC')
        my_head = xl.styles.colors.Color(rgb='00FFFF00')
        my_red_fill = xl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
        my_green_fill = xl.styles.fills.PatternFill(patternType='solid', fgColor=my_green)
        my_Header_fill = xl.styles.fills.PatternFill(patternType='solid', fgColor=my_head)
        header_font = Font(color=colors.BLACK, bold=True)
        for key, val in pk_value_idx_op_d.items():
            #Writing The Index and Key in Output Excel sheet
            sheet_obj_op.cell(val, 1, val)
            sheet_obj_op.cell(val, 2, key)
            sheet_obj_op.cell(val, 1).fill = my_green_fill
            sheet_obj_op.cell(val, 2).fill = my_green_fill
            x_tot_max_row = val

        x_tot_max_row = op_rows_idx

        print("Total rows for output sheet: " + str(x_tot_max_row) + "\n")
        print_break(80)
        print_highlight("Start processing the Data")
        ##### End Reading Primary Key Rows into New Sheet

        l_src_val = [None] * (x_tot_max_row + 1)
        l_dest_val = [None] * (x_tot_max_row + 1)


        print("Source ".ljust(25) + " = " + "Destination")
        l_regexp_for_column = 0
        for comp_len in range(0, l_xls_src_cols.__len__()):
            print(l_xls_src_cols[comp_len].ljust(25) + " = " + l_xls_dest_cols[comp_len])
            try:
                #print(f"{g_ignore_list_d[comp_len]} - {l_xls_src_cols[comp_len]}")
                l_regexp_for_column = get_ignore_case_code(g_ignore_list_d[comp_len])
            except:
                l_regexp_for_column = 0

            for x in range(1, l_max_rows_1 + 1):
                if x > l_xls_src_header_row_num:
                    #print(f"XX Key {g_src_pk_cell}")
                    cell_key_val = [sheet_obj_1.cell(row=x, column=l_src_col_nm_c_indx_d[cell_pk_nm]).value for cell_pk_nm in g_src_pk_cell]
                    l_tmp_key = str("#".join(map(str, cell_key_val)))
                    cln_key = re.sub(r"\s", "", l_tmp_key)
                    cell_val = sheet_obj_1.cell(row=x, column=l_src_col_nm_c_indx_d[l_xls_src_cols[comp_len]]).value
                    #print("ROW:" + str(pk_value_idx_op_d[cell_key_val]) +":"+str(cell_val))
                    l_src_val[(pk_value_idx_op_d[cln_key])] = cell_val

            for x in range(1, l_max_rows_2 + 1):
                if x > l_xls_dest_header_row_num:
                    #print(f"XX Key {g_dest_pk_cell}")
                    cell_key_val = [sheet_obj_2.cell(row=x, column=l_dest_col_nm_c_indx_d[cell_pk_nm]).value for cell_pk_nm in g_dest_pk_cell]
                    l_tmp_key = str("#".join(map(str, cell_key_val)))
                    cln_key = re.sub(r"\s", "", l_tmp_key)
                    cell_val = sheet_obj_2.cell(row=x, column=l_dest_col_nm_c_indx_d[l_xls_dest_cols[comp_len]]).value
                    #print("ROW:" + str(pk_value_idx_op_d[cell_key_val]) +":"+str(cell_val))
                    l_dest_val[pk_value_idx_op_d[cln_key]] = cell_val

            for l1 in range(1, l_src_val.__len__()):
                #print(f"Reg applied {l_regexp_for_column} - {l_src_val[l1]}")
                if l_regexp_for_column > 0 or isCaseignore is True:
                    l_regexp_for_column = l_regexp_for_column if (isCaseignore is False) else 7
                    l_src_regexp = ignore_Regexp(l_regexp_for_column, l_src_val[l1])
                    l_dest_regexp = ignore_Regexp(l_regexp_for_column, l_dest_val[l1])
                else:
                    l_src_regexp =  l_src_val[l1]
                    l_dest_regexp =  l_dest_val[l1]

                if l_src_regexp == l_dest_regexp:
                    l_op_cell_val = "Matching"
                    my_fill = my_green_fill
                else:
                    l_op_cell_val = "Not Matching"
                    my_fill = my_red_fill

                # Fill the data for First Row
                if l1 == 1:
                    # Fill the header with color
                    sheet_obj_op.cell(l1, op_wt_col_at + 1, l_xls_src_cols[comp_len])
                    sheet_obj_op.cell(l1, op_wt_col_at + 2, l_xls_dest_cols[comp_len])
                    sheet_obj_op.cell(l1, op_wt_col_at + 3, "Difference")

                    sheet_obj_op.cell(l1, op_wt_col_at + 1).fill = my_Header_fill
                    sheet_obj_op.cell(l1, op_wt_col_at + 2).fill = my_Header_fill
                    sheet_obj_op.cell(l1, op_wt_col_at + 3).fill = my_Header_fill

                    sheet_obj_op.cell(l1, op_wt_col_at + 1).font = header_font
                    sheet_obj_op.cell(l1, op_wt_col_at + 2).font = header_font
                    sheet_obj_op.cell(l1, op_wt_col_at + 3).font = header_font

                else:
                    # print("ROW:COL:" + str(l1) + ":" + str(op_wt_col_at + 1))
                    sheet_obj_op.cell(l1, op_wt_col_at + 1, l_src_val[l1])
                    sheet_obj_op.cell(l1, op_wt_col_at + 2, l_dest_val[l1])
                    sheet_obj_op.cell(l1, op_wt_col_at + 3, l_op_cell_val)
                    sheet_obj_op.cell(l1, op_wt_col_at + 3).fill = my_fill

                    if l_op_cell_val == 'Not Matching':
                        sheet_obj_op.cell(l1, 1).fill = my_fill
                        sheet_obj_op.cell(l1, 2).fill = my_fill

            op_wt_col_at += 3

        print_highlight("Completed processing the Data")
        l_error_flg = True
    except Exception as e:
        print_highlight(f"Error:{err.format_exc()}")
        l_error_flg = False
    else:
        wb_obj.save(p_file_nm)
        l_error_flg = True
        print_highlight("Workbook saved")

    return l_error_flg

def get_ignore_case_code(p_code = ""):
    try:
        SPACE_TAB = 1
        NEWLINE = 2
        CASE = 4
        L_CUR_VAL = 0
        l_re_exp = ""
        if p_code.__len__() == 0 or p_code == None:
            return 0

        l_buff = p_code.split("_")

        for x in l_buff:
            x_tmp = x.upper()
            if x_tmp.__eq__("SPACE") or x_tmp.__eq__("TAB"):
                L_CUR_VAL = L_CUR_VAL + SPACE_TAB
            if x_tmp.__eq__("NEWLINE"):
                L_CUR_VAL = L_CUR_VAL + NEWLINE
            if x_tmp.__eq__("CASE"):
                L_CUR_VAL = L_CUR_VAL + CASE

        # print(f"Return :{L_CUR_VAL}" )
    except:
        print(f"{err.format_exc()}")
    else:
        return L_CUR_VAL

def ignore_Regexp(p_code = 0, p_value = ""):

    try:
        if str(p_value).__len__() <= 0 or p_value == None :
            return

        l_val = p_value

        if p_code == 1: # SPACE
            l_val = re.sub(r"\s", "", str(p_value))
        elif p_code == 2: # NEWLINE
            l_val = re.sub(r"\b", "", str(p_value))
        elif p_code == 3: # SPACE NEWLINE
            l_val = re.sub(r"\s", "", str(p_value))
        elif p_code == 4: # CASE
            l_val = p_value.lower()
        elif p_code == 5: # SPACE CASE
            l_val = re.sub(r"\W", "", str(p_value)).lower()
        elif p_code == 6: # NEWLINE CASE
            l_val = re.sub(r"\b", "", str(p_value)).lower()
        elif p_code == 7: # SPACE NEWLINE CASE
            l_val = re.sub(r"\W", "", str(p_value)).lower()
        else:
            return l_val
    except:
        print(f"{err.format_exc()}")
        return l_val
    else:
        return l_val


def init(p_isSilent, isCaseignore, p_input_map_code, p_input_xls_file, p_input_map_file,
            p_MapSh, p_SrcSh, p_SrcShKey, p_DestSh, p_DestShKey, p_CompOut_Sh):
    global g_config_d

    # Read Additional Configuration File
    try:
        l_add_config_d = dict()
        if validate_file("additional.conf") == True:
            l_add_config_d = dict(rc.init("additional.conf"))
            g_config_d.update(l_add_config_d)
    except:
        print(f"No Addition Config Param loaded")
        del (l_add_config_d)

    # Reading The mapping Source and Destination Cols
    (g_src_pk_cell, g_dest_pk_cell, l_xls_src_cols, l_xls_dest_cols, l_ignore_list) = read_mapper(p_input_map_code, g_config_d, p_input_xls_file,
                                                                          p_input_map_file, p_MapSh, p_SrcSh, p_SrcShKey, p_DestSh, p_DestShKey)

    '''
        print(f"configs: {g_config_d}")
    
        print(f"init :Source Sheet:{p_SrcSh}")
        print(f"init :Source Columns:{l_xls_src_cols}")
        print(f"init :Source Key:{g_src_pk_cell}")
    
        print(f"init :Dest sheet:{p_DestSh}")
        print(f"init :Dest Columns:{l_xls_dest_cols}")
        print(f"init :Dest Key:{g_dest_pk_cell}")
    
        print(f"init :Ignore List:{l_ignore_list}")
    '''
    l_ret_val = read_xls(isCaseignore, p_input_xls_file, g_src_pk_cell, g_dest_pk_cell, l_xls_src_cols, l_xls_dest_cols, l_ignore_list, p_SrcSh, p_DestSh, p_CompOut_Sh)
    if l_ret_val == False:
        l_message = "Process Completed with Error"
    else:
        l_message = "Process Completed Successfully"

    conform_exit(p_isSilent)

if __name__ == "__main__":
    print("Main load_excel_compare")

