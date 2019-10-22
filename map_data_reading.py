import openpyxl as xl
from readConfigFile import read_config_file
import traceback as err
from global_functions import print_break

# This will read the Source and Destinatoin Mapping columns from the Column header of the source and destination sheet
def get_sd_map_col_list_from_xlSD(xl_fn, sheet1, sheet2, srcHeader=1, destHeader=1):
    # :AUTO_MAP:
    try:
        l_error_flg = False
        wb_obj = xl.load_workbook(xl_fn, data_only=True)
        # conf_sheet = p_wb_obj[g_config_d['MAP_EXCEL_SHEET_NAME']]
        # conf_sheet = p_wb_obj.worksheets[0]

        # Reading Source Column list
        if type(sheet1) == str:
            map_sh = wb_obj[sheet1]
        else:
            map_sh = wb_obj.worksheets[sheet1]

        src_col = map_sh.max_column
        print(f"Testing :{src_col}")
        src_data = list()
        print(f"Testing -{srcHeader}-{src_col}")
        for row in map_sh.iter_rows(min_row=srcHeader, min_col=1, max_row=srcHeader, max_col=src_col, values_only=True):
            print(f"Row:{row}")
            for cell in row:
                print(f"Cellvalue: {cell}")
                src_data.append(cell)

        print("Testing 2")
        # Reading Destination Column list
        if type(sheet2) == str:
            map_sh = wb_obj[sheet2]
        else:
            map_sh = wb_obj.worksheets[sheet2]

        dest_col = map_sh.max_column
        dest_data = list()

        for row in map_sh.iter_rows(min_row=destHeader, min_col=1, max_row=destHeader, max_col=dest_col, values_only=True):
            for cell in row:
                dest_data.append(cell)

        lRegexp = list()
        src_key = list()
        dest_key = list()
        l_error_flg = True
    except Exception as e:
        print(f"Error:{err.format_exc()}")
        exit(0)
    else:
        return src_key, dest_key, src_data, dest_data, lRegexp
    finally:
        wb_obj.close()
        print("Closing the File")


# This will read the Source and Destination Mapping columns from the given mapping sheet
def get_sd_map_col_list_from_xlMap(xl_fn, sheet1, rowHeader=0):
    # :EXCEL_FILE:
    try:
        l_error_flg = False
        wb_obj = xl.load_workbook(xl_fn, data_only=True)
        if type(sheet1) == str:
            map_sh = wb_obj[sheet1]
        else:
            map_sh = wb_obj.worksheets[sheet1]

        config_data = []
        conf_rows = map_sh.max_row
        conf_cols = map_sh.max_column
        # print(f"Config Sheet Rows:{conf_rows} Cols:{conf_cols}")
        for r in range(0, conf_rows):
            config_data.append([])
            for c in range(0, conf_cols):
                config_data[r].append([])
                #print(f"Map Sheet Values: row:{r} col:{c}={map_sh.cell(row=r + 1, column=c + 1).value}")
                config_data[r][c] = str(map_sh.cell(row=r + 1, column=c + 1).value)

        src_data = list()
        dest_data = list()
        lRegexp = list()
        src_key = list()
        dest_key = list()

        for k in range(0, conf_rows):
            # Skip Header row.
            if k > rowHeader:
                # print(config_data[k][0])
                kval = str(config_data[k][0]).strip()  # Key
                lval = str(config_data[k][1]).strip()  # Source
                rval = str(config_data[k][2]).strip()  # Destination
                try:
                    rReg = str(config_data[k][3]).strip()  # RegExp - Ignore List
                except:
                    rReg = ""

                #print(f"Key:{kval} Src:{lval} Dest:{rval} Reg:{rReg}")

                if (lval.__len__() <= 0 and rval.__len__() > 0) and (lval.__len__() > 0 and rval.__len__() <= 0):
                    print(f"Missing Source / destination in the Mapping excel sheet at {(k + 1)}")
                    l_error_flg = False
                    raise
                else:
                    src_data.append(lval)
                    dest_data.append(rval)
                    lRegexp.append(rReg)
                    # Getting Key Column
                    if kval.upper() == 'YES':
                        src_key.append(lval)
                        dest_key.append(rval)

        l_error_flg = True
    except Exception as e:
        print(f"Error:{err.format_exc()}")
        exit(0)
    else:
        return src_key, dest_key, src_data, dest_data, lRegexp
    finally:
        wb_obj.close()
        print("Closing the File")


# This will read the Source and Destinatoin Mapping columns from the given text file
def get_sd_map_col_list_from_txt(txt_fn):
    # :TEXT_FILE:
    l_error_flg = False
    l_config_d = dict(read_config_file(txt_fn))

    l_config_columns = 0
    src_data = list()  # list(str(g_config_d['COMP_SRC_COLS']).split(","))
    dest_data = list()  # list(str(g_config_d['COMP_DEST_COLS']).split(","))
    lRegexp = list()
    src_key = list()
    dest_key = list()
    try:
        for ls in (list(str(l_config_d['COMP_SRC_COLS']).split(","))):
            l_config_columns += 1
            src_data.append(ls.strip())

        for ls in (list(str(l_config_d['COMP_DEST_COLS']).split(","))):
            l_config_columns -= 1
            dest_data.append(ls.strip())

        for ls in (list(str(l_config_d['IGNORE_COMP_COLS']).split(","))):
            lRegexp.append(ls.strip())

        if l_config_columns != 0 or dest_data.__len__() <= 0:
            print(f"Nothing/Missing columns to compare in Config File {txt_fn}")
            l_error_flg = False
            raise

        l_xls_key_col_nm = str(l_config_d['XLSX_KEY_COL'])
        src_key = list(l_xls_key_col_nm.split(":")[0].split(','))
        dest_key = list(l_xls_key_col_nm.split(":")[1].split(','))

        l_error_flg = True
    except Exception as e:
        print(f"Error:{err.format_exc()}")
        exit(0)
    else:
        return src_key, dest_key, src_data, dest_data, lRegexp


def read_mapping_data(p_map_code, p_config_d, p_input_file, p_map_file, map_sh, src_sh, src_sh_key, dest_sh, dest_sh_key):
    # Getting the header Row
    try:
        l_xls_map_header_row_num = int(p_config_d['MAPPING_HEADER_ROW_NUM'])
    except:
        l_xls_map_header_row_num = 1

    src_key = list()
    dest_key = list()
    src_data, dest_data, lRegexp = "", "", ""

    print(f"Inputs:src_sh_key:{src_sh}-{src_sh_key}:{dest_sh}-{dest_sh_key}")

    try:
        if p_map_code == ':EXCEL_FILE:':
            # Process The Mapping Sheet with in the Excel File
            (src_key, dest_key, src_data, dest_data, lRegexp) = get_sd_map_col_list_from_xlMap(xl_fn=p_input_file, sheet1=map_sh)
        elif p_map_code == ':TEXT_FILE:':
            # Process The Mapping txt file
            (src_key, dest_key, src_data, dest_data, lRegexp) = get_sd_map_col_list_from_txt(txt_fn=p_map_file)
        elif p_map_code == ':AUTO_MAP:':
            # Process The Mapping Sheet from the Source and Destination Sheet header.
            (l_no_need_1, l_no_need_2, src_data, dest_data, lRegexp) = get_sd_map_col_list_from_xlSD(xl_fn=p_input_file, sheet1=src_sh, sheet2=dest_sh)

            print(f"src_sh_key:{src_data}-{src_sh_key}:{dest_data}-{dest_sh_key}")

            if src_sh_key.__len__() <= 0 or dest_sh_key.__len__() <= 0:
                print(f"Error: Auto Map option, must need Source and Destination Key in the parameter")
                exit(0)

            for col_l in src_sh_key:
                l_tmp = src_data[int(col_l)]
                src_key.append(l_tmp)

            for col_l in dest_sh_key:
                l_tmp = dest_data[int(col_l)]
                dest_key.append(l_tmp)

    except:
        print(f"Error in getting the Mapping data from the file")
        print(f"Error:{err.format_exc()}")
        exit(0)
    else:
        return (src_key, dest_key, src_data, dest_data, lRegexp)


def read_mapper(p_map_code, p_config_d, p_input_file, p_map_file, p_map_sh, p_src_sh, p_src_sh_key, p_dest_sh, p_dest_sh_key):
    print_break(80)
    print("Inputs")
    print(f"read_mapper: p_map_code :{p_map_code}")
    print(f"read_mapper: p_config_d : {p_config_d}")
    print(f"read_mapper: p_input_file:{p_input_file}")
    print(f"read_mapper: p_map_file: {p_map_file}")
    print(f"read_mapper: p_map_sh:{p_map_sh}")
    print(f"read_mapper: p_src_sh:{p_src_sh}")
    print(f"read_mapper: p_src_sh_key:{p_src_sh_key}")
    print(f"read_mapper: p_dest_sh: {p_dest_sh}")
    print(f"read_mapper: p_dest_sh_key:{p_dest_sh_key}")

    (src_key, dest_key, src_data, dest_data, lRegexp) = read_mapping_data(p_map_code, p_config_d, p_input_file, p_map_file, p_map_sh, p_src_sh, p_src_sh_key, p_dest_sh, p_dest_sh_key)

    print_break(80)
    print(f"read_mapper: Source Key:{src_key}")
    print(f"read_mapper: Dest Key:{dest_key}")
    print(f"read_mapper: Source Columns:{src_data}")
    print(f"read_mapper: Dest Columns:{dest_data}")
    print(f"read_mapper: Ignore List:{lRegexp}")
    print_break(80)

    return src_key, dest_key, src_data, dest_data, lRegexp


if __name__ == "__main__":
    print("Running Main Excel Map Reading")
    p_map_code = ':AUTOMAP:'
    p_config_d = {'Name':'Rajkumar'}
    #p_input_file = "Comp_Sample_AutoMap.xlsx"
    p_input_file = "Comp_Sample.xlsx"
    p_map_file = "Comp_Sample.txt"
    map_sh = "Mapping"
    src_sh = "Mapping"
    dest_sh = "MgrDest"
    src_sh_key = list()
    dest_sh_key = list()

    (src_key, dest_key, src_data, dest_data, lRegexp) = read_mapper(p_map_code, p_config_d, p_input_file, p_map_file, map_sh, src_sh, src_sh_key, dest_sh, dest_sh_key)

    print(f"Source Key:{src_key}")
    print(f"Dest Key:{dest_key}")
    print(f"Source Columns:{src_data}")
    print(f"Dest Columns:{dest_data}")
    print(f"Source Column Key:{src_sh_key}")
    print(f"Dest Column Key:{dest_sh_key}")

    print(f"Ignore List:{lRegexp}")



